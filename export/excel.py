import argparse
import json
import logging
import time
import os
from copy import copy
from collections import Counter

import openpyxl
from jsonpath_rw import parse


def cell_value(ws, row, column):
    v = ws.cell(row=row, column=column).value
    if (not v):
        return v
    return v.strip()


def read_mapping(path):
    if (not os.path.isfile(path)):
        raise Exception("No Mapping file: {}".format(path))
    if (not os.access(path, os.R_OK)):
        raise Exception("No read access to: {}".format(path))

    logging.info("Reading: {}".format(path))
    wb = openpyxl.load_workbook(path, read_only=False)
    ws = wb["key"]
    if cell_value(ws, 1, 1) != "Column":
        raise Exception('First column must be called "Column". '
                        'Worksheet "key" of file {}'.format(path))
    key_column = 1
    map_column = getColumnByName(ws, "Mapping", path)
    if not map_column:
        raise Exception("Column '{}' is not found in Worksheet key "
                        "of file {}".format("Mapping", path))
    def_column = getColumnByName(ws, "Definition", path)

    mapping = []
    for r in range(2, ws.max_row):
        cell = ws.cell(r, key_column)
        key = cell.value
        if not key:
            continue

        key = key.strip()
        style = dict()
        style["fill"] = copy(cell.fill)
        style["font"] = copy(cell.font)
        style["alignment"] = copy(cell.alignment)
        style["number_format"] = copy(cell.number_format)
        style["border"] = copy(cell.border)
        value = cell_value(ws, r, map_column)
        if value:
            def_value = cell_value(ws, r, def_column) if def_column else None
            mapping.append((len(mapping) + 1, key, value, style, def_value))

    return mapping


def read_mapping_check_tags_mapping(path):
    if (not os.path.isfile(path)):
        raise Exception("tags: No Mapping file: {}".format(path))
    if (not os.access(path, os.R_OK)):
        raise Exception("tags: No read access to: {}".format(path))

    logging.info("tags: Reading: {}".format(path))
    wb = openpyxl.load_workbook(path, read_only=False)
    ws = wb["Check_Tags"]
    key_column = 1
    res = {}
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, key_column)
        key = cell.value
        logging.info("tags: Reading key: {}".format(key))
        if not key:
            continue

        key = key.strip()
        style = dict()
        style["fill"] = copy(cell.fill)
        style["font"] = copy(cell.font)
        style["alignment"] = copy(cell.alignment)
        style["number_format"] = copy(cell.number_format)
        if key:
            res[key] = style

    return res


def getColumnByName(ws, name, path):
    for c in range(1, ws.max_column):
        if cell_value(ws, 1, c) == name:
            return c


def find_value_jsonpath(array, key):
    '''
    It's working, but very slowly
    :param array: - json
    :param key: - attribute name
    '''
    jsonpath_expr = parse('$.."{}"'.format(key))
    match = jsonpath_expr.find(array)
    if match and match[0].value:
        if isinstance(match[0].value, list):
            return ','.join([str(item) for item in match[0].value])
        else:
            return match[0].value


def find_value(array, key):
    if array.get(key):
        return array[key]
    for x in array.values():
        value = None
        if isinstance(x, dict):
            value = find_value(x, key)
        elif isinstance(x, list):
            for element in x:
                if isinstance(element, dict):
                    value = find_value(element, key)
                    if value:
                        break
        if value:
            if isinstance(value, list):
                value = ','.join([str(item) for item in value])
            return value

    return None


class ExcelExport:
    def __init__(self, template_file,
            tags_info = None, version_info = None,
            verbose_mode = False):
        self.mapping = read_mapping(template_file)
        self.check_tags_mapping = read_mapping_check_tags_mapping(
            template_file)
        self.workbook = None
        self.column_widths = {}
        self.tags_info = None
        self.tags_count = None
        self.add_tags_cfg(tags_info)
        if verbose_mode:
            for column in range(len(self.mapping)):
                print "{}: {}".format(column, self.mapping[column])
        self._new()
        if version_info:
            for idx, pair in enumerate(version_info):
                self.workbook["version"].cell(row=idx + 1, column=1,
                    value = pair[0])
                self.workbook["version"].cell(row=idx + 1, column=2,
                    value = pair[1])

    def _new(self, title=None):
        self.workbook = openpyxl.Workbook()
        ws = self.workbook.active
        ws.title = title if title else "Variants"
        for column, key, value, style, _ in self.mapping:
            if not value:
                continue
            cell = ws.cell(row=1, column=column, value=key)
            self.column_widths[cell.column] = len(key)
            for s in style:
                setattr(cell, s, style[s])

        cell = ws.cell(row=1, column=len(self.mapping) + 1, value="check tags")
        self.column_widths[cell.column] = len(cell.value)

        cell = ws.cell(row=1, column=len(self.mapping) + 2, value="tags")
        self.column_widths[cell.column] = len(cell.value)

        cell = ws.cell(row=1, column=len(self.mapping) + 3, value="tags with values")
        self.column_widths[cell.column] = len(cell.value)
        ws.freeze_panes = 'D2'
        self.__createKeySheet();

    def __createKeySheet(self):
        self.workbook.create_sheet("version")
        ws = self.workbook.create_sheet("key")
        for idx, title in enumerate(["Column", "Definition", "Mapping"]):
            ws.cell(row=1, column=idx + 1, value=title)
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = 50
        ws.freeze_panes = 'A2'
        for row, key, value, style, def_value in self.mapping:
            if not value:
                continue
            cell = ws.cell(row=row + 1, column=1, value=value)
            ws.cell(row=row + 1, column=2, value=def_value)
            ws.cell(row=row + 1, column=3, value=key)
            for s in style:
                setattr(cell, s, style[s])

    def add_tags_cfg(self, data):
        if data is None:
            return
        self.tags_info = data
        self.tags_count = Counter()

    def add_variant(self, data, tags = None):
        ws = self.workbook.active
        row = self.__get_next_row_index(ws.max_row + 1, tags)
        ws.insert_rows(row)
        for column, _, key, style, _ in self.mapping:
            if not key:
                continue
            value = self.__to_excel(find_value(data, key))
            cell = ws.cell(row=row, column=column, value=value)
            if isinstance(value, basestring):
                self.column_widths[cell.column] = max(
                    self.column_widths[cell.column], len(value))
            for s in style:
                setattr(cell, s, style[s])
        if tags is not None and self.tags_info is not None:
            self.__add_tags_to_excel(tags, row)

    def __add_tags_to_excel(self, tags, row):
        style = None
        ws = self.workbook.active
        tagList = filter(lambda k: k in self.tags_info['op-tags'], tags.keys())
        op_tags = ', '.join(tagList)
        check_tags = ', '.join(filter(lambda k: k in self.tags_info['check-tags'] and tags[k] == True, tags.keys()))
        tags_with_value = ", ".join(map(lambda t: t + ": " + tags[t].replace('\n', ' ').strip(), tagList))
        if check_tags in self.check_tags_mapping:
            style = self.check_tags_mapping[check_tags]

        col_tags = len(self.mapping) + 1
        cell = ws.cell(row=row, column = col_tags, value=check_tags)
        self.column_widths[cell.column] = max(self.column_widths[cell.column], len(cell.value))

        cell = ws.cell(row=row, column = col_tags + 1, value=op_tags)
        self.column_widths[cell.column] = max(self.column_widths[cell.column], len(cell.value))

        cell = ws.cell(row=row, column = col_tags + 2, value=tags_with_value)
        self.column_widths[cell.column] = max(self.column_widths[cell.column], len(cell.value))
        if style:
            for idx_add in range(3):
                for s in style:
                    setattr(ws.cell(row=cell.row, column=col_tags + idx_add),
                        s, style[s])

    def save(self, file):
        ws = self.workbook.active
        for column, width in self.column_widths.iteritems():
            ws.column_dimensions[column].width = min(12, width + 2)
        max_column = openpyxl.utils.get_column_letter(ws.max_column)
        ws.auto_filter.ref = 'A1:' + max_column + str(len(ws['A']))
        self.workbook.save(filename=file)

    def __to_excel(self, value):
        if isinstance(value, str) and value.startswith("http"):
            return '=HYPERLINK("{0}","{0}")'.format(value)
        if isinstance(value, dict):
            return '=HYPERLINK("{}","{}")'.format(value["link"], value["title"])
        return value

    def __get_next_row_index(self, max_index, tags):
        if tags is None:
            return max_index
        check_tag = next(iter(filter(lambda k:
            k in self.tags_info['check-tags'] and tags[k] == True,
            tags.keys())), None)
        if check_tag is None:
            return max_index

        res = 0
        for key in self.tags_info['check-tags']:
            res += self.tags_count[key]
            if check_tag == key:
                self.tags_count[key] += 1
                return min(res + 2, max_index)

        return max_index;

if __name__ == '__main__':
    import Enum
    class LoadMode(Enum):
        RECORD = "@RECORD",
        TAGS_CFG = "@TAGS_CFG",
        TAGS = "@TAGS",


    def processing(args):
        start_time = time.time()
        print "parsing template {} ...".format(args.template)
        export = ExcelExport(args.template, verbose_mode = args.verbose)
        print "export variants from {} ...".format(args.input)
        with open(args.input) as json_file:
            mode = LoadMode.RECORD
            record = None
            for idx, line in enumerate(json_file):
                if line.startswith("@"):
                    mode = LoadMode[line.strip()[1:]]
                else:
                    data = json.loads(line)
                    if mode == LoadMode.RECORD:
                        record = data
                    elif mode == LoadMode.TAGS_CFG:
                        export.add_tags_cfg(data)
                    elif mode == LoadMode.TAGS:
                        if record != None:
                            export.add_variant(record, data)
                        record = None

                if args.limit and idx >= args.limit:
                    break
                if args.verbose and idx > 0 and idx % 100 == 0:
                    print "export lines: {}".format(idx)

            print "total export line: {}".format(idx)

        print "save {}".format(args.output)
        export.save(args.output)
        print "complete (execution time: {0:.3f} s)".format(time.time() - start_time)


    parser = argparse.ArgumentParser()
    parser.add_argument("-t", "--template", help="template file", required=True)
    parser.add_argument("-i", "--input", help="input file with json lines", required=True)
    parser.add_argument("-o", "--output", help="result file name", required=True)
    parser.add_argument("-l", "--limit", help="maximum number of rows to export", type=int)
    parser.add_argument("-v", "--verbose", help="increase output verbosity", action="store_true")
    args = parser.parse_args()
    processing(args)
